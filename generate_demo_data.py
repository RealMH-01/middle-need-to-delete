#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
演示数据生成脚本。

一键生成脱敏的模拟公司文件夹结构 + 模板文件目录，用于面试 / 演示。

用法：
    python generate_demo_data.py                 # 默认生成到 项目根目录/demo_company/
    python generate_demo_data.py /path/to/dir    # 生成到指定目录

生成后：
    1. 目标目录下有完整的 <订单根文件夹>/ 树（业务员 → 中间层 → 客户 → 订单号）；
    2. 目标目录下有 模板文件/ 目录，含各个产地+文档类型的空模板文件；
    3. 目标目录下有 .order_tool/ 目录，预填好 config.json / salespersons.json /
       history.json 以及 templates/ 下的标准模板；
    4. ~/.order_tool_bootstrap.json 会被更新为指向此目录，运行 python main.py
       即可直接进入"已配置好"的状态。

实现说明：
    订单号文件夹的内部结构直接调用 app/core/folder_builder.execute_build，
    复用生产逻辑，避免重复实现。模板文件仅创建空文件（xlsx 通过 openpyxl
    生成只含表头的 workbook，doc / xls 等创建 0 字节占位文件即可；脚本不
    引入任何新依赖——openpyxl 已经在 requirements.txt 中）。
"""

from __future__ import annotations

import json
import os
import sys
from pathlib import Path
from typing import Dict, List

# 让脚本独立运行时能 import 到项目内的 app 包
_ROOT = Path(__file__).resolve().parent
if str(_ROOT) not in sys.path:
    sys.path.insert(0, str(_ROOT))

# 延迟导入（路径加好之后）
from app.core import folder_builder  # noqa: E402
from app.core import storage as storage_mod  # noqa: E402
from app.core.storage import Storage  # noqa: E402


# ------------------------------------------------------------------
# 演示数据定义
# ------------------------------------------------------------------
# 一名业务员（可能有分公司前缀）→ 其客户列表；
# 每个客户 → 订单号列表，订单号前缀决定订单类型（EXP=外贸 / DOM=内贸）。
DEMO_SALESPERSONS: List[Dict] = [
    {
        "name": "张三",
        "rel_path": "张三",
        "mid_layer": "1.进行订单",
        "customers": {
            "示例客户A": ["EXP-2026001", "EXP-2026002"],
            "示例客户B": ["DOM-2026003"],
        },
    },
    {
        "name": "李四",
        "rel_path": "李四",
        "mid_layer": "1.进行订单",
        "customers": {
            "示例客户C": ["EXP-2026010", "EXP-2026011"],
            "示例客户D": ["EXP-2026012", "DOM-2026013"],
            "示例客户E": ["DOM-2026014", "DOM-2026015", "DOM-2026016"],
        },
    },
    {
        "name": "王五",
        "rel_path": "王五",
        "mid_layer": "",
        "customers": {
            "示例客户F": ["EXP-2026020", "EXP-2026021"],
            "示例客户G": ["DOM-2026022"],
        },
    },
    {
        "name": "赵六",
        "rel_path": "华南分公司/赵六",
        "mid_layer": "",
        "customers": {
            "示例客户H": ["EXP-2026030", "DOM-2026031"],
            "示例客户I": ["EXP-2026032"],
        },
    },
    {
        "name": "孙七",
        "rel_path": "华南分公司/孙七",
        "mid_layer": "",
        "customers": {
            "示例客户J": ["EXP-2026040"],
            "示例客户K": ["DOM-2026041", "DOM-2026042"],
        },
    },
]


# 模板文件目录下需要生成的空模板文件（路径相对 模板文件/）
DEMO_TEMPLATE_FILES: List[str] = [
    "通用/CG.xlsx",
    "外贸通用/CI.xlsx",
    "外贸通用/PL.xls",
    "外贸通用/托书.doc",
    "华北工厂/华北工厂外贸生产.doc",
    "华北工厂/华北工厂外贸发货.docx",
    "华北工厂/华北工厂内贸生产.xlsx",
    "华北工厂/华北工厂内贸发货.xlsx",
    "华南工厂/华南工厂外贸生产.xlsx",
    "华南工厂/华南工厂外贸发货.xlsx",
]


# 产品类别池（按 config.json 的默认 origin_map 选）
_PRODUCT_POOL = ["环氧树脂", "其他产品"]


# ------------------------------------------------------------------
# 工具：生成空模板文件
# ------------------------------------------------------------------
def _write_empty_xlsx(path: Path) -> None:
    """用 openpyxl 生成一个只含表头的空 xlsx。"""
    try:
        from openpyxl import Workbook
    except ImportError:  # pragma: no cover
        # openpyxl 没装的极端情况：直接写空文件占位
        path.write_bytes(b"")
        return
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "示例模板"
    ws["A2"] = "（此处由实际业务填充）"
    wb.save(str(path))


def _write_empty_placeholder(path: Path) -> None:
    """创建 0 字节的占位文件（用于 .doc / .docx / .xls 等）。"""
    path.write_bytes(b"")


def _write_template_file(target_file: Path) -> None:
    target_file.parent.mkdir(parents=True, exist_ok=True)
    suffix = target_file.suffix.lower()
    if suffix == ".xlsx":
        _write_empty_xlsx(target_file)
    else:
        _write_empty_placeholder(target_file)


# ------------------------------------------------------------------
# 生成流程
# ------------------------------------------------------------------
def _product_category_for(order_no: str, idx: int) -> str:
    """按订单号决定产品类别，让演示数据多样化"""
    return _PRODUCT_POOL[idx % len(_PRODUCT_POOL)]


def _order_type_for(order_no: str) -> str:
    return "外贸" if order_no.startswith("EXP") else "内贸"


def _template_file_for(order_type: str, storage: Storage):
    """选择对应订单类型的标准模板"""
    fn = Storage.standard_template_filename(order_type)
    tmpl = storage.load_template(fn)
    return fn, tmpl


def _create_order_folders(storage: Storage, template_files_dir: Path,
                          cfg: Dict) -> int:
    """按 DEMO_SALESPERSONS 定义创建所有订单的文件夹结构。

    返回创建的订单数量。
    """
    origin_map = cfg.get("origin_map", {})
    origin_file_ext = cfg.get("origin_file_ext", {})
    total_orders = 0
    order_idx = 0

    for sp in DEMO_SALESPERSONS:
        sp_name = sp["name"]
        for customer, orders in sp["customers"].items():
            for order_no in orders:
                order_type = _order_type_for(order_no)
                product_category = _product_category_for(order_no, order_idx)
                order_idx += 1

                _, tmpl = _template_file_for(order_type, storage)
                if not tmpl:
                    print(f"[WARN] 未找到 {order_type} 标准模板，跳过 {order_no}")
                    continue

                # 客户目录：<root>/<订单根>/<rel_path>/[mid_layer/]<客户>/
                base_path = storage.build_customer_dir(sp_name, customer)
                os.makedirs(base_path, exist_ok=True)

                order = {
                    "order_no": order_no,
                    "customer": customer,
                    "po_no": f"PO-DEMO-{order_no}",
                    "product_info": "示例产品 200KG",
                    "salesperson": sp_name,
                    "custom_no": "",
                    "order_type": order_type,
                    "product_category": product_category,
                    "needs_inspection": False,
                }
                try:
                    folder_builder.execute_build(
                        order=order,
                        template=tmpl,
                        base_path=base_path,
                        template_files_dir=str(template_files_dir),
                        origin_map=origin_map,
                        origin_file_ext=origin_file_ext,
                    )
                    total_orders += 1
                except Exception as e:
                    print(f"[WARN] 创建 {sp_name}/{customer}/{order_no} 失败：{e}")

    return total_orders


def _write_salespersons(storage: Storage) -> None:
    """把演示业务员写入 salespersons.json（去掉"customers 是 dict"的结构）。"""
    items = []
    for sp in DEMO_SALESPERSONS:
        items.append({
            "name": sp["name"],
            "rel_path": sp["rel_path"],
            "mid_layer": sp["mid_layer"],
            "customers": list(sp["customers"].keys()),
        })
    storage.save_salespersons(items)


def _write_template_files(template_files_dir: Path) -> None:
    """在 模板文件/ 目录下生成演示所需的 10 个模板文件。"""
    template_files_dir.mkdir(parents=True, exist_ok=True)
    for rel in DEMO_TEMPLATE_FILES:
        _write_template_file(template_files_dir / rel)


def _update_bootstrap(target_dir: Path) -> None:
    bootstrap_path = storage_mod.BOOTSTRAP_FILE
    bootstrap_path.parent.mkdir(parents=True, exist_ok=True)
    data = {"last_root": str(target_dir)}
    with open(bootstrap_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ------------------------------------------------------------------
# 主入口
# ------------------------------------------------------------------
def generate(target_dir: Path) -> None:
    target_dir = target_dir.resolve()
    print(f"\n📁 目标目录：{target_dir}")

    if target_dir.exists() and any(target_dir.iterdir()):
        print(f"⚠️  目标目录已存在且非空。将在其中继续生成/补齐演示数据。")
    target_dir.mkdir(parents=True, exist_ok=True)

    # 1. 初始化 Storage（会自动创建 .order_tool / templates / standard_*.json
    #    以及把 origin_map / origin_file_ext / order_root_folder / mid_layer_keywords
    #    四个默认字段写入 config.json）
    print("🔧 初始化 .order_tool/ 并写入默认模板与配置 ...")
    storage = Storage(str(target_dir))

    # 2. 生成模板文件目录 + 空模板文件
    template_files_dir = target_dir / "模板文件"
    print(f"📄 生成模板文件目录：{template_files_dir}")
    _write_template_files(template_files_dir)

    # 把 template_files_dir 写进 config.json，让程序启动后能自动复制模板
    storage.update_config(template_files_dir=str(template_files_dir))

    # 3. 写入演示业务员（不含客户→订单的映射，仅 name / rel_path / mid_layer / customers 列表）
    print(f"👥 写入 {len(DEMO_SALESPERSONS)} 个业务员到 salespersons.json ...")
    _write_salespersons(storage)

    # 4. 按演示数据创建订单文件夹结构
    cfg = storage.load_config()
    print("📦 按模板创建订单文件夹 ...")
    total_orders = _create_order_folders(storage, template_files_dir, cfg)

    # 5. 把 bootstrap 指向此目录
    print("🧭 更新 ~/.order_tool_bootstrap.json 指向此演示目录 ...")
    _update_bootstrap(target_dir)

    # 6. 完成提示
    total_salespersons = len(DEMO_SALESPERSONS)
    order_root = storage.order_root_folder
    print("\n" + "=" * 60)
    print("✅ 演示数据生成完成！")
    print("=" * 60)
    print(f"   目标目录        ：{target_dir}")
    print(f"   订单根文件夹    ：{target_dir / order_root}")
    print(f"   模板文件目录    ：{template_files_dir}")
    print(f"   .order_tool     ：{target_dir / '.order_tool'}")
    print(f"   业务员数量      ：{total_salespersons}")
    print(f"   订单数量        ：{total_orders}")
    print(f"   bootstrap       ：{storage_mod.BOOTSTRAP_FILE}")
    print("=" * 60)
    print("🚀 现在运行 `python main.py` 即可体验完整功能。")
    print("   程序会自动加载演示数据（根目录、模板目录、业务员等均已配置好）。")
    print("=" * 60 + "\n")


def main() -> None:
    if len(sys.argv) >= 2:
        target = Path(sys.argv[1]).expanduser()
    else:
        target = _ROOT / "demo_company"
    generate(target)


if __name__ == "__main__":
    main()
