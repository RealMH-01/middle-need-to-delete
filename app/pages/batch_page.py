# -*- coding: utf-8 -*-
"""批量导入页 —— Neo-brutalism 视觉适配 + Delegate 替换 setCellWidget。

改造要点：
- 「订单类型」「产品类别」两列改用 QStyledItemDelegate，
  平时显示纯文字，双击/点击编辑时才弹出下拉框，选完即消失；
- 其余业务逻辑、Excel 导入/导出、黄色警告标记等保持不变。
"""

import os
from datetime import datetime

from PyQt5.QtCore import Qt, pyqtSignal
from PyQt5.QtGui import QBrush, QColor
from PyQt5.QtWidgets import (
    QCheckBox, QComboBox, QFileDialog, QGridLayout, QGroupBox, QHBoxLayout,
    QHeaderView, QLabel, QLineEdit, QMessageBox, QPushButton, QSpinBox,
    QStyledItemDelegate, QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget
)

from ..core import folder_builder
from ..style import COLOR_SECONDARY
from ..widgets import StyledComboBox


HEADERS = ["订单类型", "订单号", "客户名称", "产品信息", "客户PO号", "产品类别", "是否商检", "业务员", "状态"]

_WARN_BG = QColor(COLOR_SECONDARY)
_WARN_FG = QColor("#000000")


# =====================================================================
# Delegate：平时纯文字，编辑时弹出下拉框
# =====================================================================
class _ComboDelegate(QStyledItemDelegate):
    """通用的下拉框委托。

    Parameters
    ----------
    items : list[str]
        下拉选项列表，可通过 :meth:`set_items` 动态更新。
    parent : QWidget, optional
    """

    def __init__(self, items=None, parent=None):
        super().__init__(parent)
        self._items = list(items or [])

    def set_items(self, items):
        self._items = list(items or [])

    def createEditor(self, parent, option, index):
        editor = QComboBox(parent)
        editor.addItems(self._items)
        # 选中后立即提交并关闭，无需再点别处
        editor.activated.connect(lambda: self.commitData.emit(editor))
        editor.activated.connect(lambda: self.closeEditor.emit(editor))
        return editor

    def setEditorData(self, editor, index):
        value = index.data(Qt.DisplayRole) or ""
        idx = editor.findText(value)
        if idx >= 0:
            editor.setCurrentIndex(idx)
        else:
            editor.setCurrentIndex(0)

    def setModelData(self, editor, model, index):
        model.setData(index, editor.currentText(), Qt.EditRole)

    def updateEditorGeometry(self, editor, option, index):
        editor.setGeometry(option.rect)


class BatchPage(QWidget):
    request_back = pyqtSignal()
    request_help = pyqtSignal(str)

    def __init__(self, storage, parent=None):
        super().__init__(parent)
        self.storage = storage
        self._build_ui()

    # ============== UI ==============
    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 14, 20, 14)
        root.setSpacing(10)

        # ------ 顶部栏 ------
        top = QHBoxLayout()
        btn_back = QPushButton("← 返回首页")
        btn_back.setObjectName("SecondaryButton")
        btn_back.clicked.connect(self.request_back.emit)
        top.addWidget(btn_back)
        title = QLabel("批量导入")
        title.setObjectName("TitleLabel")
        top.addWidget(title)
        top.addStretch(1)

        btn_help = QPushButton("❓ 帮助")
        btn_help.setObjectName("SecondaryButton")
        btn_help.setToolTip("打开右侧帮助面板，查看批量导入的详细说明")
        btn_help.clicked.connect(lambda: self.request_help.emit("sec-batch"))
        top.addWidget(btn_help)

        root.addLayout(top)

        # ----- 身份与模板 -----
        id_group = QGroupBox("① 身份与模板（批量订单共用）")
        id_layout = QGridLayout(id_group)
        id_layout.addWidget(QLabel("业务员："), 0, 0)
        self.cmb_sales = StyledComboBox(searchable=True)
        self.cmb_sales.setMinimumWidth(180)
        id_layout.addWidget(self.cmb_sales, 0, 1)
        id_layout.addWidget(QLabel("客户（可留空，使用每行的客户名称）："), 0, 2)
        self.cmb_customer = StyledComboBox(searchable=True)
        self.cmb_customer.setMinimumWidth(200)
        self.cmb_customer.setInsertPolicy(QComboBox.InsertAtBottom)
        id_layout.addWidget(self.cmb_customer, 0, 3)
        tip = QLabel(
            "提示：批量导入时，模板按「订单类型 + 业务员 + 客户」自动匹配；"
            "若 Excel 中「业务员」列在系统中不存在，对应行会被"
            "<span style='background:#FFD93D;color:#000;padding:0 4px;'>"
            "黄色警告</span>标记。"
        )
        tip.setStyleSheet("color:#000000;")
        tip.setWordWrap(True)
        tip.setTextFormat(Qt.RichText)
        id_layout.addWidget(tip, 1, 0, 1, 4)
        self.cmb_sales.currentIndexChanged.connect(self._reload_customers)
        root.addWidget(id_group)

        # ----- 导入方式 -----
        op_group = QGroupBox("② 导入订单")
        op_layout = QHBoxLayout(op_group)
        btn_dl = QPushButton("下载 Excel 模板")
        btn_dl.setObjectName("SecondaryButton")
        btn_dl.clicked.connect(self._download_template)
        btn_import = QPushButton("导入 Excel")
        btn_import.clicked.connect(self._import_excel)
        btn_add = QPushButton("＋ 添加一行")
        btn_add.setObjectName("SecondaryButton")
        btn_add.clicked.connect(lambda: self._add_row())
        btn_del = QPushButton("－ 删除选中行")
        btn_del.setObjectName("SecondaryButton")
        btn_del.clicked.connect(self._del_rows)
        self.spin_rows = QSpinBox()
        self.spin_rows.setRange(1, 200)
        self.spin_rows.setValue(5)
        btn_gen = QPushButton("按预设行数生成")
        btn_gen.setObjectName("SecondaryButton")
        btn_gen.clicked.connect(self._gen_rows)

        op_layout.addWidget(btn_dl)
        op_layout.addWidget(btn_import)
        op_layout.addSpacing(20)
        op_layout.addWidget(btn_add)
        op_layout.addWidget(btn_del)
        op_layout.addSpacing(20)
        op_layout.addWidget(QLabel("预设行数："))
        op_layout.addWidget(self.spin_rows)
        op_layout.addWidget(btn_gen)
        op_layout.addStretch(1)
        root.addWidget(op_group)

        # ----- 表格 -----
        self.table = QTableWidget(0, len(HEADERS) + 1)
        all_headers = ["序号"] + HEADERS
        self.table.setHorizontalHeaderLabels(all_headers)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table.setColumnWidth(0, 50)
        self.table.setColumnWidth(1, 80)
        self.table.setColumnWidth(2, 180)
        self.table.setColumnWidth(3, 180)
        self.table.setColumnWidth(4, 160)
        self.table.setColumnWidth(5, 140)
        self.table.setColumnWidth(6, 100)
        self.table.setColumnWidth(7, 80)
        self.table.setColumnWidth(8, 120)
        self.table.setColumnWidth(9, 220)

        # ★ 用 Delegate 替代 setCellWidget
        self._type_delegate = _ComboDelegate(["外贸", "内贸"], self.table)
        self.table.setItemDelegateForColumn(1, self._type_delegate)

        self._cat_delegate = _ComboDelegate([], self.table)
        self.table.setItemDelegateForColumn(6, self._cat_delegate)

        root.addWidget(self.table, 1)

        # ----- 执行 -----
        bottom = QHBoxLayout()
        bottom.addStretch(1)
        btn_preview = QPushButton("预览全部")
        btn_preview.setObjectName("SecondaryButton")
        btn_preview.clicked.connect(self._preview_all)
        btn_run = QPushButton("确认批量创建")
        btn_run.setStyleSheet("font-size:14px; font-weight:bold; padding:8px 24px;")
        btn_run.clicked.connect(self._run_all)
        bottom.addWidget(btn_preview)
        bottom.addWidget(btn_run)
        root.addLayout(bottom)

    # ============== 外部入口 ==============
    def refresh(self):
        self._load_salespersons()
        self._refresh_category_column_visibility()

    # ============== 数据 ==============
    def _load_salespersons(self):
        self.cmb_sales.blockSignals(True)
        self.cmb_sales.clear()
        names = [it["name"] for it in self.storage.load_salespersons()]
        self.cmb_sales.addItems([""] + names)
        self.cmb_sales.blockSignals(False)
        self._reload_customers()

    def _reload_customers(self):
        sales = self.cmb_sales.currentText()
        self.cmb_customer.clear()
        if sales:
            self.cmb_customer.addItems([""] + self.storage.get_customers(sales))
        else:
            self.cmb_customer.addItems([""])

    def _refresh_category_column_visibility(self):
        cfg = self.storage.load_config() if self.storage else {}
        om = cfg.get("origin_map") or {}
        has_cats = bool(om)
        self.table.setColumnHidden(6, not has_cats)
        # ★ 动态更新产品类别 Delegate 的选项列表
        self._cat_delegate.set_items(list(om.keys()))

    # ============== 表格操作 ==============
    def _add_row(self, data=None):
        r = self.table.rowCount()
        self.table.insertRow(r)

        # 序号
        it = QTableWidgetItem(str(r + 1))
        it.setFlags(it.flags() & ~Qt.ItemIsEditable)
        it.setTextAlignment(Qt.AlignCenter)
        self.table.setItem(r, 0, it)

        # ★ 订单类型：纯文字 item，编辑时由 Delegate 弹出下拉
        order_type = "外贸"
        if data and data.get("order_type") in ("外贸", "内贸"):
            order_type = data["order_type"]
        self.table.setItem(r, 1, QTableWidgetItem(order_type))

        # 订单号
        self.table.setItem(
            r, 2, QTableWidgetItem(data.get("order_no", "") if data else ""))
        # 客户
        self.table.setItem(
            r, 3, QTableWidgetItem(data.get("customer", "") if data else ""))
        # 产品信息
        self.table.setItem(
            r, 4, QTableWidgetItem(data.get("product_info", "") if data else ""))
        # 客户PO号
        self.table.setItem(
            r, 5, QTableWidgetItem(data.get("po_no", "") if data else ""))

        # ★ 产品类别：纯文字 item，编辑时由 Delegate 弹出下拉
        cfg = self.storage.load_config() if self.storage else {}
        category_options = list((cfg.get("origin_map") or {}).keys())
        cat_value = ""
        if data and data.get("product_category") in category_options:
            cat_value = data["product_category"]
        elif category_options:
            cat_value = category_options[0]
        self.table.setItem(r, 6, QTableWidgetItem(cat_value))

        # 是否商检
        chk = QCheckBox()
        w = QWidget()
        lay = QHBoxLayout(w)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setAlignment(Qt.AlignCenter)
        lay.addWidget(chk)
        if data and data.get("needs_inspection"):
            chk.setChecked(True)
        self.table.setCellWidget(r, 7, w)

        # 业务员
        per_row_sp = (data.get("salesperson") or "").strip() if data else ""
        self.table.setItem(r, 8, QTableWidgetItem(per_row_sp))

        # 状态
        self.table.setItem(r, 9, QTableWidgetItem(""))

        self._renumber()

    def _del_rows(self):
        rows = sorted({i.row() for i in self.table.selectedIndexes()}, reverse=True)
        if not rows:
            QMessageBox.information(self, "提示", "请选中要删除的行")
            return
        for r in rows:
            self.table.removeRow(r)
        self._renumber()

    def _gen_rows(self):
        self.table.setRowCount(0)
        for _ in range(self.spin_rows.value()):
            self._add_row()

    def _renumber(self):
        for r in range(self.table.rowCount()):
            it = self.table.item(r, 0)
            if it:
                it.setText(str(r + 1))

    # ============== Excel ==============
    def _download_template(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "保存 Excel 模板", "批量导入模板.xlsx", "Excel 文件 (*.xlsx)")
        if not path:
            return
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = "批量导入"
        headers = ["订单类型", "订单号", "客户名称", "产品信息", "客户PO号",
                   "产品类别", "是否需要商检", "业务员"]
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2196F3")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        cfg_sample = self.storage.load_config() if self.storage else {}
        sample_category = ""
        om_sample = cfg_sample.get("origin_map") or {}
        if om_sample:
            sample_category = next(iter(om_sample.keys()))
        ws.append(["外贸", "EXP-2026001", "示例客户A", "产品示例 200KG",
                   "PO-2026-001", sample_category, "是", "示例业务员"])
        ws.append(["内贸", "DOM-2026002", "示例客户B", "产品示例 1T",
                   "", sample_category, "否", ""])

        type_comment = Comment("请填写「外贸」或「内贸」", "订单文件夹工具")
        ws.cell(row=1, column=1).comment = type_comment

        cat_list = list(om_sample.keys())
        if cat_list:
            cat_str = "、".join(cat_list)
            default_hint = (
                f"留空则使用默认值「{sample_category}」。"
                if sample_category else "留空则跳过 [产地] 模板文件的复制。"
            )
            comment_text = (
                f"请填写以下产品类别之一：\n{cat_str}\n\n{default_hint}"
            )
        else:
            comment_text = (
                "当前未配置产品类别。如需使用此功能，"
                "请在程序首页「⚙ 高级设置」中配置产地映射。"
            )
        ws.cell(row=1, column=6).comment = Comment(
            comment_text, "订单文件夹工具"
        )

        insp_comment = Comment(
            "请填写「是」或「否」，仅对外贸订单有效", "订单文件夹工具"
        )
        ws.cell(row=1, column=7).comment = insp_comment

        try:
            sp_names = [it["name"] for it in self.storage.load_salespersons()]
        except Exception:
            sp_names = []
        if sp_names:
            sp_str = "、".join(sp_names[:20])
            if len(sp_names) > 20:
                sp_str += f"…共{len(sp_names)}人"
            sp_comment_text = (
                "可选。不填则使用页面顶部选中的业务员。\n\n"
                f"系统中已有的业务员：\n{sp_str}"
            )
        else:
            sp_comment_text = "可选。不填则使用页面顶部选中的业务员。"
        ws.cell(row=1, column=8).comment = Comment(
            sp_comment_text, "订单文件夹工具"
        )

        widths = [12, 22, 24, 28, 18, 12, 14, 12]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        wb.save(path)
        QMessageBox.information(self, "成功", f"模板已保存到：\n{path}")

    def _import_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if not path:
            return
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                QMessageBox.warning(self, "提示", "Excel 为空")
                return
            header = [str(x).strip() if x else "" for x in rows[0]]
            def idx(name):
                for i, h in enumerate(header):
                    if name in h:
                        return i
                return -1
            i_type = idx("订单类型")
            i_no = idx("订单号")
            i_cust = idx("客户名称")
            i_prod = idx("产品信息")
            i_po = idx("PO号")
            i_cat = idx("产品类别")
            i_insp = idx("商检")
            i_sp = idx("业务员")

            cfg_import = self.storage.load_config() if self.storage else {}
            valid_categories = list((cfg_import.get("origin_map") or {}).keys())
            default_category = valid_categories[0] if valid_categories else ""

            known_salespersons = {
                it["name"] for it in self.storage.load_salespersons()
            }
            invalid_rows = []
            invalid_names = []

            start_row = self.table.rowCount()

            count = 0
            for r in rows[1:]:
                if not r or all(x is None or str(x).strip() == "" for x in r):
                    continue
                data = {
                    "order_type": (str(r[i_type]).strip() if i_type >= 0 and r[i_type] else "外贸"),
                    "order_no": (str(r[i_no]).strip() if i_no >= 0 and r[i_no] else ""),
                    "customer": (str(r[i_cust]).strip() if i_cust >= 0 and r[i_cust] else ""),
                    "product_info": (str(r[i_prod]).strip() if i_prod >= 0 and r[i_prod] else ""),
                    "po_no": (str(r[i_po]).strip() if i_po >= 0 and r[i_po] else ""),
                    "product_category": (str(r[i_cat]).strip() if i_cat >= 0 and r[i_cat] else default_category),
                    "salesperson": (str(r[i_sp]).strip() if i_sp >= 0 and r[i_sp] else ""),
                }
                if valid_categories and data["product_category"] not in valid_categories:
                    data["product_category"] = default_category
                if data["order_type"] not in ("外贸", "内贸"):
                    data["order_type"] = "外贸"
                insp_raw = str(r[i_insp]).strip() if i_insp >= 0 and r[i_insp] is not None else ""
                data["needs_inspection"] = insp_raw in ("是", "Y", "y", "YES", "yes", "1", "true", "True", "✓")

                per_row_sp = data["salesperson"].strip()
                row_invalid = bool(per_row_sp) and per_row_sp not in known_salespersons

                self._add_row(data)
                new_row_idx = self.table.rowCount() - 1
                if row_invalid:
                    invalid_rows.append(new_row_idx)
                    invalid_names.append(per_row_sp)
                    self._mark_row_warning(
                        new_row_idx,
                        f"⚠ 业务员「{per_row_sp}」未在系统中找到，请先新增或用扫描导入",
                    )
                count += 1

            if invalid_rows:
                uniq_names = sorted(set(invalid_names))
                preview = "、".join(uniq_names[:10])
                if len(uniq_names) > 10:
                    preview += "…"
                QMessageBox.warning(
                    self, "导入完成（含警告）",
                    f"已导入 {count} 行，其中 {len(invalid_rows)} 行业务员在系统中未找到：\n"
                    f"{preview}\n\n"
                    "这些行已用<黄色>背景高亮，可在执行前到首页"
                    "「🧭 扫描导入业务员」或手动添加业务员后再批量创建。"
                )
            else:
                QMessageBox.information(self, "成功", f"已导入 {count} 行。")
        except Exception as e:
            QMessageBox.critical(self, "错误", f"解析 Excel 失败：{e}")

    def _mark_row_warning(self, row: int, status_text: str = ""):
        for col in range(self.table.columnCount()):
            item = self.table.item(row, col)
            if item is not None:
                item.setBackground(QBrush(_WARN_BG))
                item.setForeground(QBrush(_WARN_FG))
        if status_text:
            it = QTableWidgetItem(status_text)
            it.setBackground(QBrush(_WARN_BG))
            it.setForeground(QBrush(_WARN_FG))
            self.table.setItem(row, 9, it)

    # ============== 采集 ==============
    def _collect_rows(self):
        rows = []
        common_sales = self.cmb_sales.currentText().strip()

        if common_sales and self.cmb_sales.findText(common_sales) < 0:
            QMessageBox.warning(
                self, "提示",
                f"业务员「{common_sales}」不在列表中，请从下拉列表中选择。"
            )
            return []

        common_customer = self.cmb_customer.currentText().strip()

        cat_col_hidden = self.table.isColumnHidden(6)
        for r in range(self.table.rowCount()):
            # ★ 改为从 item 读取，不再用 cellWidget
            type_item = self.table.item(r, 1)
            order_type = type_item.text().strip() if type_item else "外贸"
            if order_type not in ("外贸", "内贸"):
                order_type = "外贸"

            no_item = self.table.item(r, 2)
            order_no = no_item.text().strip() if no_item else ""
            sp_item = self.table.item(r, 8)
            per_row_sp = (sp_item.text().strip() if sp_item else "")
            customer = self.table.item(r, 3).text().strip() if self.table.item(r, 3) else ""
            product_info = self.table.item(r, 4).text().strip() if self.table.item(r, 4) else ""
            po_no = self.table.item(r, 5).text().strip() if self.table.item(r, 5) else ""

            if cat_col_hidden:
                product_category = ""
            else:
                # ★ 改为从 item 读取
                cat_item = self.table.item(r, 6)
                product_category = cat_item.text().strip() if cat_item else ""
                # 无效值静默置空
                cfg = self.storage.load_config() if self.storage else {}
                valid_cats = list((cfg.get("origin_map") or {}).keys())
                if product_category and valid_cats and product_category not in valid_cats:
                    product_category = ""

            chk_w = self.table.cellWidget(r, 7)
            chk = chk_w.findChild(QCheckBox) if chk_w else None
            needs_inspection = bool(chk and chk.isChecked())
            if not order_no:
                continue
            rows.append({
                "row_index": r,
                "order_type": order_type,
                "order_no": order_no,
                "customer": customer or common_customer,
                "product_info": product_info,
                "po_no": po_no,
                "product_category": product_category,
                "needs_inspection": needs_inspection and order_type == "外贸",
                "salesperson": (per_row_sp or common_sales).strip(),
            })
        return rows

    def _set_status(self, row, text, color="#333"):
        it = QTableWidgetItem(text)
        it.setForeground(QBrush(QColor(color)))
        self.table.setItem(row, 9, it)

    # ============== 预览 & 执行 ==============
    def _preview_all(self):
        rows = self._collect_rows()
        if not rows:
            QMessageBox.information(self, "提示", "没有有效订单行")
            return
        for od in rows:
            try:
                base_path = self.storage.build_customer_dir(
                    od["salesperson"], od["customer"])
                if not base_path:
                    self._set_status(od["row_index"], "路径无效", "#E53935")
                    continue
                _, tpl = self.storage.match_template(od["salesperson"], od["customer"], od["order_type"])
                if not tpl:
                    self._set_status(od["row_index"], "无可用模板", "#E53935")
                    continue
                ctx = folder_builder.build_context(od)
                tpl_folders = folder_builder.flatten_template_folders(
                    tpl, base_path, ctx, od["needs_inspection"])
                tpl_folders, _ = folder_builder.compare_with_existing(
                    base_path, tpl_folders)
                exists = sum(1 for i in tpl_folders if i["status"] == "existing" and not i.get("is_root"))
                tocre = sum(1 for i in tpl_folders if i["status"] == "to_create" and not i.get("is_root"))
                root_existed = any(i["status"] == "existing" and i.get("is_root") for i in tpl_folders)
                if root_existed:
                    self._set_status(od["row_index"], f"待补建 {tocre}，已存在 {exists}", "#1976D2")
                else:
                    self._set_status(od["row_index"], f"待创建 {tocre} 个", "#1976D2")
            except Exception as e:
                self._set_status(od["row_index"], f"错误：{e}", "#E53935")

    def _run_all(self):
        rows = self._collect_rows()
        if not rows:
            QMessageBox.information(self, "提示", "没有有效订单行")
            return
        reply = QMessageBox.question(self, "确认",
                                     f"即将批量创建 {len(rows)} 笔订单文件夹，是否继续？")
        if reply != QMessageBox.Yes:
            return
        cfg = self.storage.load_config()
        tpl_dir = cfg.get("template_files_dir") or None
        success, fail = 0, 0
        details = []
        for od in rows:
            try:
                base_path = self.storage.build_customer_dir(
                    od["salesperson"], od["customer"])
                if not base_path:
                    self._set_status(od["row_index"], "路径无效", "#E53935")
                    fail += 1
                    continue
                tpl_name, tpl = self.storage.match_template(od["salesperson"], od["customer"], od["order_type"])
                if not tpl:
                    self._set_status(od["row_index"], "无可用模板", "#E53935")
                    fail += 1
                    continue
                result = folder_builder.execute_build(
                    order=od, template=tpl, base_path=base_path,
                    template_files_dir=tpl_dir,
                    origin_map=cfg.get("origin_map") or {},
                    origin_file_ext=cfg.get("origin_file_ext") or {})
                slim_copy_results = []
                for cr in result["copy_results"]:
                    src = cr.get("src", "") or ""
                    dst = cr.get("dst", "") or ""
                    slim_copy_results.append({
                        "copied": bool(cr.get("copied")),
                        "src": os.path.basename(src) if src else "",
                        "dst": os.path.basename(dst) if dst else "",
                        "reason": cr.get("reason", ""),
                    })
                self.storage.append_history({
                    "time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "operator": cfg.get("operator", ""),
                    "salesperson": od["salesperson"],
                    "customer": od["customer"],
                    "order_no": od["order_no"],
                    "order_type": od["order_type"],
                    "product_category": od["product_category"],
                    "template_name": tpl_name,
                    "path": result["base_path"],
                    "result": "成功",
                    "created_count": len(result["created"]),
                    "skipped_count": len(result["skipped"]),
                    "copied_count": sum(1 for r in result["copy_results"] if r.get("copied")),
                    "detail": {
                        "created": list(result["created"]),
                        "skipped": list(result["skipped"]),
                        "copy_results": slim_copy_results,
                        "checklist_path": result.get("checklist_path", ""),
                    },
                })
                self._set_status(od["row_index"],
                                 f"✅ 新建 {len(result['created'])}，跳过 {len(result['skipped'])}，复制模板 {sum(1 for r in result['copy_results'] if r.get('copied'))}",
                                 "#2E7D32")
                success += 1
                details.append(f"[{od['order_no']}] 新建 {len(result['created'])} / 跳过 {len(result['skipped'])} / 复制 {sum(1 for r in result['copy_results'] if r.get('copied'))}")
            except Exception as e:
                self._set_status(od["row_index"], f"❌ {e}", "#E53935")
                fail += 1
                details.append(f"[{od['order_no']}] 失败：{e}")

        from PyQt5.QtWidgets import QDialog, QPlainTextEdit, QPushButton, QVBoxLayout
        dlg = QDialog(self)
        dlg.setWindowTitle("批量执行完成")
        dlg.resize(700, 460)
        v = QVBoxLayout(dlg)
        v.addWidget(QLabel(f"<b>成功 {success} 笔，失败 {fail} 笔。</b>"))
        tx = QPlainTextEdit()
        tx.setReadOnly(True)
        tx.setPlainText("\n".join(details))
        v.addWidget(tx)
        btn = QPushButton("关闭")
        btn.clicked.connect(dlg.accept)
        v.addWidget(btn)
        dlg.exec_()
